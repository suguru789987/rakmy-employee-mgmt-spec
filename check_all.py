#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
従業員管理 引継ぎ資料の整合性チェック

使い方:  python3 check_all.py
資料を直したら毎回これを流す。NG が 0 件なら、資料どうしの食い違いは無い。

対象:
  仕様書 / 受入条件表（定義・確認表）/ 検証プラン / 検証データセット / ヘルプページ / 使い方
  それぞれの TSV・Excel・Markdown と、デスクトップの配布物
"""
import csv, re, os, glob, math, collections, sys

try:
    import openpyxl
except ImportError:
    print('openpyxl が必要です: pip3 install openpyxl'); sys.exit(1)

ROOT = os.path.dirname(os.path.abspath(__file__))
os.chdir(ROOT)
DESK = os.path.expanduser('~/Desktop/従業員管理_20260804/')

F_SPEC = '20260803_01_従業員管理_仕様書_PdM版.md'
F_AC   = '20260803_02_従業員管理_受入条件表.tsv'
F_CHK  = '20260803_02b_従業員管理_受入条件_確認表.tsv'
F_PLAN = '20260803_03_従業員管理_検証プラン.tsv'
F_HOW  = '20260803_04_受入条件表と確認表の使い方.md'
F_HELP = '20260805_ヘルプページ_従業員管理_マニュアル版.md'
X_DEF  = '20260805_従業員管理_00_定義表.xlsx'
X_CHK  = '20260805_従業員管理_01_確認表.xlsx'
X_PLAN = '20260805_従業員管理_02_検証プラン.xlsx'
X_DATA = '20260805_従業員管理_04_検証データセット.xlsx'
X_HELP = '20260805_従業員管理_05_ヘルプページ.xlsx'
X_HOW  = '20260805_従業員管理_使い方.xlsx'

RE_TID = r'\b[SCTR]-[A-Z]?\d+\b'          # S-01 T-D04 C-008 R-02
RE_AC  = r'\bAC-\d+\b'

results = []          # (区分, 見出し, ok, 詳細)


def check(cat, name, ok, detail=''):
    results.append((cat, name, bool(ok), detail))


def T(p):
    with open(p, encoding='utf-8-sig') as f:
        return list(csv.DictReader(f, delimiter='\t'))


def D(p):
    d = '\t' if p.endswith('.tsv') else ','
    with open(p, encoding='utf-8-sig') as f:
        return list(csv.DictReader(f, delimiter=d))


def lines_needed(text, width):
    if not text:
        return 1
    n = 0
    for seg in str(text).split('\n'):
        w = sum(2 if ord(c) > 0x7f else 1 for c in seg)
        n += max(1, math.ceil(w / max(4, width - 2)))
    return n


# ============================================================ 1. ファイル
def c1_files():
    need = [F_SPEC, F_AC, F_CHK, F_PLAN, F_HOW, F_HELP, X_DEF, X_CHK, X_PLAN, X_DATA, X_HELP, X_HOW]
    miss = [f for f in need if not os.path.exists(f)]
    check('1 ファイル', '必要な資料が揃っている', not miss, f'欠落: {miss}' if miss else f'{len(need)}件')

    files = sorted(glob.glob('data/*'))
    nums = sorted(int(os.path.basename(f)[:2]) for f in files)
    check('1 ファイル', f'データセットが{len(files)}ファイルで連番',
          nums == list(range(1, len(files) + 1)), f'{len(files)}ファイル / 欠番{sorted(set(range(1,len(files)+1))-set(nums))}')

    nobom = []
    for p in glob.glob('*.tsv') + glob.glob('data/*'):
        with open(p, 'rb') as f:
            if f.read(3) != b'\xef\xbb\xbf':
                nobom.append(os.path.basename(p))
    check('1 ファイル', 'TSV/CSVがBOM付き（Excelで文字化けしない）', not nobom, f'BOM無し: {nobom}' if nobom else '全22件')

    if os.path.isdir(DESK):
        import hashlib
        h = lambda p: hashlib.md5(open(p, 'rb').read()).hexdigest()
        diff = [f for f in need if os.path.exists(DESK + f) and h(f) != h(DESK + f)]
        absent = [f for f in need if not os.path.exists(DESK + f)]
        dd = [f for f in os.listdir('data')
              if not os.path.exists(DESK + 'data/' + f) or h('data/' + f) != h(DESK + 'data/' + f)]
        check('1 ファイル', 'デスクトップと内容が一致', not (diff or absent or dd),
              f'差分:{diff} 未配置:{absent} data:{dd}' if (diff or absent or dd) else '13件＋data19件')
    else:
        check('1 ファイル', 'デスクトップの配布先が存在', False, DESK)


# ============================================================ 2. 受入条件
def c2_acceptance():
    ac, chk = T(F_AC), T(F_CHK)
    ws = openpyxl.load_workbook(X_CHK).active
    TOP = 13
    BOT = TOP + len(ac) - 1
    xl = [ws.cell(r, 4).value for r in range(TOP, BOT + 1) if ws.cell(r, 4).value]

    check('2 受入条件', f'件数が3資料で一致（{len(ac)}件）',
          len(ac) == len(chk) == len(xl), f'定義{len(ac)} 確認表{len(chk)} Excel{len(xl)}')

    s1, s2, s3 = {a['条件ID'] for a in ac}, {a['条件ID'] for a in chk}, set(xl)
    check('2 受入条件', '条件IDの集合が一致', s1 == s2 == s3, f'差分 {(s1 ^ s2) | (s1 ^ s3)}')

    m = {a['条件ID']: a for a in ac}
    bad = []
    for r in range(TOP, BOT + 1):
        i = ws.cell(r, 4).value
        if not i:
            continue
        for col, key in [(5, '判定レベル'), (6, '満たすべき状態'), (7, '合格基準（定量）'), (10, '対象画面')]:
            if ws.cell(r, col).value != m[i][key]:
                bad.append(f'{i}:{key}')
    check('2 受入条件', 'ExcelとTSVの中身が一致', not bad, f'{bad[:5]}' if bad else '51条件×4項目')

    lv = collections.Counter(a['判定レベル'] for a in ac)
    check('2 受入条件', '判定レベルがL0/L1/L2に収まり合計が一致',
          set(lv) <= {'L0 リリース不可', 'L1 MVP必須', 'L2 次段階'} and sum(lv.values()) == len(ac), dict(lv))

    # Excel の集計式が合否列を指しているか
    f5 = str(ws.cell(5, 6).value or '')
    hd = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(12, c).value == '合否':
            hd = openpyxl.utils.get_column_letter(c)
    check('2 受入条件', '集計式が合否列と全条件の範囲を指している',
          hd and f'${hd}${TOP}:${hd}${BOT}' in f5, f'合否={hd} 想定={TOP}〜{BOT} 式={f5[:70]}')

    dv = [str(d.sqref) for d in ws.data_validations.dataValidation]
    check('2 受入条件', '合否がOK/NG/対象外のプルダウン', any(hd and hd in x for x in dv), dv)


# ============================================================ 3. 検証プラン
def c3_plan():
    tp = T(F_PLAN)
    ws = openpyxl.load_workbook(X_PLAN).active
    prep = sum(1 for r in tp if r['段階'] == '準備')
    check('3 検証プラン', f'TSVとExcelの行数が一致（準備{prep}＋検証{len(tp)-prep}）',
          ws.max_row - 2 == len(tp), f'TSV{len(tp)} Excel{ws.max_row-2}')
    check('3 検証プラン', '1シート構成', len(openpyxl.load_workbook(X_PLAN).sheetnames) == 1,
          openpyxl.load_workbook(X_PLAN).sheetnames)

    need_cols = ['段階', '検証ID', '検証内容', '操作・前提条件', '使用する検証データ', '期待値',
                 'スクショファイル名', 'スクショ（撮るもの）', 'スクショ貼付欄', '画面', '確認ポイント',
                 'ヘルプページ該当箇所', '対応受入条件', '実際の値', '判定', '所見・気づき', '実施日', '実施者', '計算根拠']
    miss = [c for c in need_cols if c not in tp[0]]
    check('3 検証プラン', '必要な列が揃っている', not miss, f'欠落 {miss}' if miss else f'{len(tp[0])}列')

    ops = [r for r in tp if r['段階'] != '準備']
    rules = [
        ('期待値に確認場所が明示', lambda r: r['期待値'].startswith('【確認場所')),
        ('期待値が判定可能', lambda r: bool(re.search(r'\d|=0|されない|される|一致|変わらない|残らない', r['期待値']))),
        ('画面が具体的（18字以上）', lambda r: len(r['画面']) >= 18),
        ('確認ポイントがある', lambda r: len(r['確認ポイント']) >= 6),
        ('使用する検証データが記入', lambda r: len(r['使用する検証データ']) >= 5),
        ('スクショの撮るものが記入', lambda r: len(r['スクショ（撮るもの）']) >= 8),
        ('受入条件が紐づく', lambda r: bool(re.search(RE_AC, r['対応受入条件'])) or r['対応受入条件'].startswith('（') or r['対応受入条件'] == '-'),
        ('ヘルプページが紐づく', lambda r: '＞' in r['ヘルプページ該当箇所'] or r['ヘルプページ該当箇所'].startswith('（')),
        ('計算根拠がある', lambda r: len(r['計算根拠']) >= 6),
    ]
    for name, f in rules:
        bad = [r['検証ID'] for r in ops if not f(r)]
        check('3 検証プラン', name, not bad, f'{len(ops)-len(bad)}/{len(ops)} 不足:{bad[:5]}')


# ============================================================ 4. 紐付け
def c4_links():
    tp, chk, ac = T(F_PLAN), T(F_CHK), T(F_AC)
    ids = {r['検証ID'] for r in tp}
    acid = {a['条件ID'] for a in ac}

    bad = [(r['検証ID'], x) for r in tp for x in re.findall(RE_AC, r['対応受入条件']) if x not in acid]
    check('4 紐付け', '検証プラン→受入条件の参照が実在', not bad, bad[:5])

    bad = [(k['条件ID'], x) for k in chk for x in re.findall(RE_TID, k['検証ID']) if x not in ids]
    check('4 紐付け', '確認表→検証プランの参照が実在', not bad, bad[:5])

    fwd = collections.defaultdict(set)
    for r in tp:
        for a in re.findall(RE_AC, r['対応受入条件']):
            fwd[a].add(r['検証ID'])
    bwd = {k['条件ID']: set(re.findall(RE_TID, k['検証ID'])) for k in chk}
    diff = [a for a in acid if fwd.get(a, set()) != bwd.get(a, set())]
    check('4 紐付け', '双方向の紐付けが一致', not diff,
          f'不一致 {[(a, sorted(fwd.get(a,set())), sorted(bwd.get(a,set()))) for a in diff[:3]]}')

    uncovered = sorted(acid - set(fwd))
    l2only = all(next(a['判定レベル'] for a in ac if a['条件ID'] == x).startswith('L2') for x in uncovered) if uncovered else True
    check('4 紐付け', '検証が無い条件はL2のみ', l2only, f'{uncovered}')

    # ヘルプ見出しの実在
    hp = open(F_HELP, encoding='utf-8').read()
    heads = set(re.findall(r'^### (.+)$', hp, re.M))
    bad = []
    for r in tp:
        v = r['ヘルプページ該当箇所']
        if not v or v.startswith('（'):
            continue
        for part in [x.strip() for x in v.split('／') if x.strip()]:
            sec = part.split('＞')[-1].strip()
            if sec not in heads:
                bad.append((r['検証ID'], part))
    check('4 紐付け', '検証プランのヘルプ参照が実在', not bad, bad[:5])

    ws = openpyxl.load_workbook(X_CHK).active
    hcol = next((c for c in range(1, ws.max_column + 1) if ws.cell(12, c).value == 'ヘルプページ該当箇所'), None)
    bad = []
    if hcol:
        for r in range(13, 13 + len(ac)):
            v = str(ws.cell(r, hcol).value or '')
            for part in [x.strip() for x in v.split('／') if x.strip()]:
                sec = part.split('＞')[-1].strip()
                if sec and sec not in heads:
                    bad.append((ws.cell(r, 4).value, part))
    check('4 紐付け', '受入条件Excelのヘルプ参照が実在', hcol and not bad, bad[:5] if bad else f'列{hcol}')

    # 実装レビュー
    need = [a['条件ID'] for a in ac if a['確認方法'] == '操作＋実装レビュー']
    m = {k['条件ID']: k['検証ID'] for k in chk}
    miss = [x for x in need if not re.search(r'R-\d+', m.get(x, ''))]
    check('4 紐付け', '実装レビュー13条件にR-が紐づく', not miss, f'{len(need)}件中 不足{miss}')


# ============================================================ 5. 実行順
def c5_order():
    tp = T(F_PLAN)
    o = {r['検証ID']: i for i, r in enumerate(tp)}
    for x in ['T-S04', 'T-D07']:
        check('5 実行順', f'{x} が締め済みの区間内（T-S03〜T-S05）',
              o['T-S03'] < o[x] < o['T-S05'], f"S03={o['T-S03']} {x}={o[x]} S05={o['T-S05']}")
    check('5 実行順', 'T-A08（勤怠取込）が締めより前', o['T-A08'] < o['T-S03'])
    calc = [r['検証ID'] for r in tp if r['段階'] == '計算']
    check('5 実行順', 'C-000が計算の先頭', calc[0] == 'C-000', calc[:3])
    check('5 実行順', 'C-008が計算の最後（5月の金額を変えるため）', calc[-1] == 'C-008', calc[-3:])
    check('5 実行順', '移行検証の後にC-000で初期化', o['C-000'] > o['T-Z03'])

    # 削除対象がマスタのみの従業員か
    used = collections.defaultdict(set)
    for f in glob.glob('data/*'):
        if os.path.basename(f).startswith(('07_', '08_', '19_', '20_')):
            continue
        for r in D(f):
            for k in ('従業員コード', 'スタッフコード'):
                if r.get(k):
                    used[r[k]].add(os.path.basename(f)[:2])
    safe = {e for e, s in used.items() if s == {'02'}}
    dels = [(r['検証ID'], re.findall(r'EMP-\d+', r['操作・前提条件']))
            for r in tp if '削除' in r['操作・前提条件'] and r['段階'] != '準備' and r['検証ID'].startswith(('T-L', 'T-D'))]
    bad = [(i, e) for i, es in dels for e in es if e not in safe]
    check('5 実行順', '削除対象がマスタのみの従業員', not bad, f'安全={sorted(safe)} 指定={dels}')


# ============================================================ 6. データセット
def c6_data():
    emp = {r['従業員コード']: r for r in D('data/02_従業員マスタ.tsv')}
    shops = {r['店舗ID'] for r in D('data/01_店舗マスタ.tsv')}
    m = D('data/05_月次給与実績.tsv')

    # 参照整合
    ext = {r['外部ID'] for r in D('data/20_外部ID対応表.tsv')}
    bad = []
    for f in glob.glob('data/*'):
        if os.path.basename(f).startswith(('07_', '08_', '19_', '20_')):
            continue          # 取込CSVは外部ID体系。対応表で解決する
        for i, r in enumerate(D(f), 2):
            c = r.get('従業員コード') or r.get('スタッフコード')
            if c and c not in emp and not re.match(r'EMP-(013|014|015|099)$', c):
                bad.append(f'{os.path.basename(f)}:{i}:{c}')
    check('6 データ', '従業員コードがマスタに実在', not bad, bad[:5])

    codes = {r['従業員コード'] for r in D('data/20_外部ID対応表.tsv')}
    bad = sorted(codes - set(emp))
    check('6 データ', '外部ID対応表の従業員コードがマスタに実在', not bad, bad[:5])
    noid = [r['従業員コード'] for r in D('data/02_従業員マスタ.tsv') if not r.get('従業員ID')]
    check('6 データ', '全従業員に従業員IDが付番されている', not noid, noid[:5])

    # 日次→月次
    agg = collections.defaultdict(lambda: {'h': 0.0, 'help': 0.0, 'n': 0.0})
    for r in D('data/13_勤務データ_2026-05_全稼働日.tsv'):
        a = agg[r['従業員コード']]
        a['h'] += float(r['実働時間']); a['help'] += float(r['ヘルプ時間'] or 0); a['n'] += float(r['深夜残業時間'] or 0)
    bad = []
    for r in m:
        a = agg.get(r['従業員コード'])
        if not a:
            bad.append(f"{r['従業員コード']}:日次なし"); continue
        for col, k in [('総労働時間', 'h'), ('ヘルプ時間', 'help'), ('深夜残業時間', 'n')]:
            if abs(float(r[col]) - a[k]) > 0.01:
                bad.append(f"{r['従業員コード']}:{col}")
    check('6 データ', 'data/13の日次合計＝data/05の月次', not bad, bad[:5] or f'{len(m)}名')

    g = collections.Counter()
    for r in D('data/06_日次給与実績.tsv'):
        g[r['従業員コード']] += int(r['日次人件費'])
    bad = [r['従業員コード'] for r in m if g[r['従業員コード']] != int(r['実績給与'])]
    check('6 データ', 'data/06の日次人件費合計＝実績給与', not bad, bad or f'全社{sum(g.values()):,}円')

    # 算式（判断14：交通費を人件費に含める）
    helpfare = collections.defaultdict(dict)
    for r in D('data/03_ヘルプ先交通費.tsv'):
        helpfare[r['従業員コード']][r['ヘルプ先店舗']] = int(r['交通費'])
    days = collections.defaultdict(lambda: {'work': 0, 'help': collections.Counter()})
    for r in D('data/06_日次給与実績.tsv'):
        a = days[r['従業員コード']]
        a['work'] += 1
        if r['ヘルプ先'] != '-':
            a['help'][r['ヘルプ先']] += 1
    bad = []
    for r in m:
        k = r['従業員コード']; e = emp[k]
        base = int(e['単位給与額']) / float(e['所定労働時間']) if e['給与単位'] == '月給' else float(e['単位給与額'])
        shop_fare = int(e['単位交通費']) if e['交通費単位'] == '月額' else int(e['単位交通費']) * days[k]['work']
        help_fare = sum(helpfare[k].get(s, 0) * n for s, n in days[k]['help'].items())
        # 判断23：所属店舗の交通費は概算給与に含まれる
        want = int(r['概算給与']) + int(r['深夜残業代']) + int(r['みなし超過残業代']) + help_fare
        if int(r['実績給与']) != want:
            bad.append(f"{k}:実績給与({r['実績給与']}≠{want})")
        if float(r['総労働時間']) > 0:
            # 按分の基礎は交通費を除いた給与部分（判断14・23）
            pay = int(r['概算給与']) - shop_fare + int(r['深夜残業代']) + int(r['みなし超過残業代'])
            hp = round(pay * float(r['ヘルプ時間']) / float(r['総労働時間'])) + help_fare
            if abs(int(r['ヘルプ人件費']) - hp) > 2:
                bad.append(f"{k}:ヘルプ人件費({r['ヘルプ人件費']}≠{hp})")
        if abs(float(r['深夜残業時間'])) > 0:
            if abs(int(r['深夜残業代']) - int(base * 1.5 * float(r['深夜残業時間']))) > 1:
                bad.append(f'{k}:深夜残業代')
    check('6 データ', '確定した算式で全行を再現できる（交通費込み）', not bad, bad[:5])

    # 検証プランの金額が data/05 と矛盾しないか
    vals = set()
    for r in m:
        for c in ['概算給与', '深夜残業代', 'みなし超過残業代', '実績給与', 'ヘルプ人件費']:
            vals.add(f'{int(r[c]):,}')
    vals.add(f"{sum(int(r['実績給与']) for r in m):,}")
    KNOWN = {'357,600', '371,886', '334,286', '340,000', '360,000', '410,000', '290,000', '332,727', '352,286', '334,286', '147,000', '350,000', '320,000', '380,000',
             '280,000', '300,000', '144,000', '312,768', '332,388', '129,000', '1,561,312'}
    bad = []
    for r in T(F_PLAN):
        for num in set(re.findall(r'\d{1,3}(?:,\d{3})+', r['期待値'] + r['操作・前提条件'])):
            if 100000 <= int(num.replace(',', '')) <= 2000000 and num not in vals and num not in KNOWN:
                bad.append(f"{r['検証ID']}:{num}")
    check('6 データ', '検証プランの金額がdata/05と整合', not bad, bad[:5])

    # データセットExcel
    wb = openpyxl.load_workbook(X_DATA); ws = wb['00_目次']
    bad = []
    for r in range(5, 30):
        name = ws.cell(r, 1).value
        if not name:
            break
        real = len(D('data/' + name))
        sn = name.rsplit('.', 1)[0][:31]
        srow = wb[sn].max_row - 3 if sn in wb.sheetnames else -1
        if not (int(ws.cell(r, 2).value) == real == srow):
            bad.append(f'{name}:目次{ws.cell(r,2).value}/実{real}/シート{srow}')
    check('6 データ', 'データセットExcelの行数が実データと一致', not bad, bad[:5] or '19ファイル')


# ============================================================ 7. ヘルプ
def c7_help():
    hp = open(F_HELP, encoding='utf-8').read()
    h2 = re.findall(r'^## (.+)$', hp, re.M)
    check('7 ヘルプ', '大項目が 目次／操作手順／ケース別／よくあるご質問',
          h2 == ['目次', '操作手順', 'ケース別', 'よくあるご質問'], h2)

    heads = re.findall(r'^### (.+)$', hp, re.M)
    toc = re.search(r'## 目次\n\n\|[\s\S]*?\n\n', hp)
    check('7 ヘルプ', '目次がある', bool(toc))

    FLOW = {1: ['設定インポート'], 2: ['従業員登録'], 3: ['予約'], 4: ['設定変更履歴一覧'],
            5: ['月別実績'], 6: ['締め'], 7: ['締め済み期間の翌日以降'], 8: ['設定・実績エクスポート'],
            9: ['設定取込データ'], 10: ['実績データ取込'], 11: ['従業員設定']}
    miss = [n for n, ks in FLOW.items() if not any(k in hp for k in ks)]
    check('7 ヘルプ', '仕様書の11フローを網羅', not miss, f'不足フロー{miss}')

    shots = re.findall(r'^📸 (.+)$', hp, re.M)
    check('7 ヘルプ', '撮影位置が明示されている', len(shots) >= 10, f'{len(shots)}箇所')

    NG = ['会社スコープの機能です', 'BOM付き', 'バッジ濃色', '移動元']
    hit = [w for w in NG if w in hp.split('-->')[1]]
    check('7 ヘルプ', '社内用語・仕様専用の記述が残っていない', not hit, hit)

    check('7 ヘルプ', 'ASCII罫線の図が無い', not re.search(r'[├└│─┌┐┘┬┴┼]', hp))


# ============================================================ 8. 件数表記
def c8_counts():
    ac, tp = T(F_AC), T(F_PLAN)
    n_data = len(glob.glob('data/*'))
    how = open(F_HOW, encoding='utf-8').read()
    spec = open(F_SPEC, encoding='utf-8').read()

    check('8 件数表記', f'使い方が data/ {n_data}ファイル と書いている',
          f'{n_data}ファイル' in how, re.findall(r'`data/` \d+ファイル', how))
    check('8 件数表記', f'仕様書が {n_data}ファイル と書いている',
          f'{n_data}ファイル' in spec, re.findall(r'\*\*\d+ファイル\*\*', spec))

    ws = openpyxl.load_workbook(X_DEF).active
    txt = '\n'.join(str(ws.cell(r, c).value or '') for r in range(1, 60) for c in range(1, 5))
    check('8 件数表記', f'定義表が data/01〜{n_data} と書いている', f'01〜{n_data}' in txt,
          re.findall(r'data/01〜\d+', txt))

    prep = sum(1 for r in tp if r['段階'] == '準備')
    ws2 = openpyxl.load_workbook(X_PLAN).active
    note = str(ws2.cell(1, 1).value or '')
    check('8 件数表記', f'検証プランの注記が 準備{prep}行＋検証{len(tp)-prep}ケース',
          f'準備{prep}行' in note and f'検証{len(tp)-prep}ケース' in note, note[:40])

    lv = collections.Counter(a['判定レベル'] for a in ac)
    for k, v in lv.items():
        lab = k.split()[0]
        check('8 件数表記', f'使い方が {lab}={v}件 と書いている', f'{v}件' in how, k)

    if os.path.exists('README.md'):
        rd = open('README.md', encoding='utf-8').read()
        old = re.findall(r'(\d+)条件', rd)
        bad = [x for x in old if int(x) != len(ac)]
        check('8 件数表記', f'READMEが {len(ac)}条件 と書いている', not bad, f'古い記載 {bad}')

    # --- リリース判断・使い方に書いた件数（受入条件表を足せば必ず動く数字）---
    n_l0 = lv['L0 リリース不可']
    n_l1 = lv['L1 MVP必須']
    n_stop = sum(1 for a in ac
                 if a['判定レベル'].startswith('L1') and '不可' in a['不合格のときリリースは'])
    check('8 件数表記', f'仕様書のGo条件が L0（{n_l0}件）', f'L0（{n_l0}件）' in spec,
          re.search(r'L0（\d+件）', spec).group(0) if re.search(r'L0（\d+件）', spec) else '記載なし')
    check('8 件数表記', f'仕様書のGo条件が L1（{n_l1}件）', f'L1（{n_l1}件）' in spec,
          re.search(r'L1（\d+件）', spec).group(0) if re.search(r'L1（\d+件）', spec) else '記載なし')
    check('8 件数表記', f'仕様書のGo条件が 不可{n_stop}件', f'{n_stop}件が全合格' in spec,
          re.search(r'\d+件が全合格', spec).group(0) if re.search(r'\d+件が全合格', spec) else '記載なし')
    check('8 件数表記', f'使い方のNo-Go条件が L0（{n_l0}件）', f'L0（{n_l0}件）' in how,
          re.search(r'L0（\d+件）', how).group(0) if re.search(r'L0（\d+件）', how) else '記載なし')

    n_rv = sum(1 for a in ac if a['確認方法'] == '操作＋実装レビュー')
    check('8 件数表記', f'使い方が 操作＋実装レビューの{n_rv}条件 と書いている',
          f'実装レビュー」の{n_rv}条件' in how,
          re.search(r'実装レビュー」の(\d+)条件', how).group(0) if re.search(r'実装レビュー」の\d+条件', how) else '記載なし')

    # --- 判断の件数（判断記録を足せば必ず動く数字）---
    if os.path.exists('JUDGMENT_LOG.md'):
        jl = open('JUDGMENT_LOG.md', encoding='utf-8').read()
        n_fix = jl.count('✅ **確定**')
        n_open = jl.count('❌ 未確定')
        check('8 件数表記', f'仕様書のR6が 判断{n_fix + n_open}件（確定{n_fix}・未確定{n_open}）',
              f'判断{n_fix + n_open}件（確定{n_fix}・未確定{n_open}）' in spec,
              re.search(r'判断\d+件[^が]*が代替', spec).group(0)[:40] if re.search(r'判断\d+件', spec) else '記載なし')


# ============================================================ 9. 体裁
def c9_layout():
    for path, cols in [(X_PLAN, None), (X_CHK, None)]:
        ws = openpyxl.load_workbook(path).active
        bad = 0
        hrow = 2 if path == X_PLAN else 12
        for r in range(hrow + 1, ws.max_row + 1):
            avail = int((ws.row_dimensions[r].height or 15) // 15)
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if not cell.alignment or not cell.alignment.wrap_text:
                    continue
                w = ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width or 10
                if lines_needed(cell.value, w) > avail:
                    bad += 1
        check('9 体裁', f'{os.path.basename(path)} で文字が潰れていない', not bad, f'{bad}箇所')


# ============================================================ 10. ID・連番
def c10_ids():
    ac, tp, chk = T(F_AC), T(F_PLAN), T(F_CHK)

    n = [int(a['条件ID'].split('-')[1]) for a in ac]
    dup = [k for k, v in collections.Counter(n).items() if v > 1]
    check('10 ID・連番', '受入条件IDに重複が無い', not dup, f'重複 {dup}')
    gap = sorted(set(range(min(n), max(n) + 1)) - set(n))
    check('10 ID・連番', f'受入条件IDが連番（AC-{min(n)}〜AC-{max(n)}・{len(n)}件）', not gap, f'欠番 {gap}')

    ids = [r['検証ID'] for r in tp]
    dup = [k for k, v in collections.Counter(ids).items() if v > 1]
    check('10 ID・連番', '検証IDに重複が無い', not dup, f'重複 {dup}')

    dupc = [k for k, v in collections.Counter(a['条件ID'] for a in chk).items() if v > 1]
    check('10 ID・連番', '確認表の条件IDに重複が無い', not dupc, f'重複 {dupc}')

    # 検証IDの欠番は「統合で空いた番号」。参照が壊れるので詰めない。件数だけ記録する
    miss = {}
    for pre in ['S', 'C', 'R', 'T-L', 'T-R', 'T-D', 'T-H', 'T-A', 'T-S', 'T-I', 'T-E', 'T-N', 'T-Z']:
        g = sorted(int(x.rsplit('-', 1)[1]) for x in ids if x.startswith(pre + '-') and x.rsplit('-', 1)[1].isdigit())
        if g:
            gp = sorted(set(range(min(g), max(g) + 1)) - set(g))
            if gp:
                miss[pre] = gp
    check('10 ID・連番', '検証IDの欠番が使い方に説明されている',
          ('欠番' in open(F_HOW, encoding='utf-8').read()) or not miss,
          f'欠番あり {miss}（統合で空いた番号。詰めると参照が壊れる）')

    dn = sorted(int(os.path.basename(f)[:2]) for f in glob.glob('data/*'))
    gap = sorted(set(range(min(dn), max(dn) + 1)) - set(dn))
    dup = [k for k, v in collections.Counter(dn).items() if v > 1]
    check('10 ID・連番', f'データセットが連番（{min(dn):02d}〜{max(dn):02d}）', not gap and not dup, f'欠番{gap} 重複{dup}')

    bad = [(r['検証ID'], x) for r in tp for x in re.findall(r'\bS-\d+\b', r['初期設定（前提データ）']) if x not in set(ids)]
    check('10 ID・連番', '前提に書いた準備IDが実在', not bad, bad[:5])

    bad = [(r['検証ID'], r['スクショファイル名']) for r in tp
           if r['スクショファイル名'] and r['スクショファイル名'] != r['検証ID'] + '.png']
    check('10 ID・連番', 'スクショ名が「検証ID.png」', not bad, bad[:5])

    bad = [f"data/{x}" for r in tp for x in re.findall(r'data/(\d{2})', r['操作・前提条件'] + r['使用する検証データ'])
           if not glob.glob(f'data/{x}_*')]
    check('10 ID・連番', '本文が参照するdataファイルが実在', not bad, sorted(set(bad))[:5])


# ============================================================ 11. 資料間の言い回し
def c11_wording():
    spec = open(F_SPEC, encoding='utf-8').read()
    help_ = open(F_HELP, encoding='utf-8').read()
    tp = T(F_PLAN)
    plan_txt = '\n'.join(r['計算根拠'] + r['期待値'] for r in tp)

    FORMULA = [
        ('基礎時給', ['単位給与額 ÷ 所定労働時間', '単位給与額÷所定労働時間']),
        ('実績給与', ['概算給与 ＋ 深夜残業代 ＋ みなし超過残業代', '概算給与＋深夜残業代＋みなし超過残業代']),
        ('想定給与に所属店舗の交通費を含む', ['所属店舗の交通費', '＋ 所属店舗の交通費']),
        ('ヘルプ人件費', ['給与部分 × ヘルプ時間 ÷ 総労働時間', '給与部分×ヘルプ時間÷総労働時間']),
        ('実績給与にヘルプ先の交通費を含む', ['みなし超過残業代 ＋ ヘルプ先の交通費', 'みなし超過残業代 ＋ **ヘルプ先の交通費']),
    ]
    for name, pats in FORMULA:
        ok_s = any(p in spec for p in pats)
        ok_h = any(p in help_ for p in pats)
        check('11 言い回し', f'{name}の算式が仕様書とヘルプで一致', ok_s and ok_h, f'仕様書{ok_s} ヘルプ{ok_h}')

    check('11 言い回し', '深夜割増1.5が3資料で一致',
          '1.5' in spec and '1.5' in help_ and '1.5' in plan_txt)

    # 暫定の判断がヘルプの掲載前チェックに残っているか
    note = help_.split('-->')[0]
    check('11 言い回し', 'ヘルプの掲載前チェックに判断7・8の確定日が明記',
          '判断7' in note and '判断8' in note and '2026-08-07' in note, note[:60])
    check('11 言い回し', '検証プランに判断7の確定が書かれている', '判断7（2026-08-07 確定）' in plan_txt)

    # 対象月
    check('11 言い回し', '検証の対象月が2026-05で揃っている',
          sum(1 for r in tp if '2026-05' in r['操作・前提条件']) >= 5)


# ============================================================ 12. 派生ファイル
def c12_derived():
    md = open(F_HELP, encoding='utf-8').read()
    body = re.sub(r'<!--[\s\S]*?-->\n*', '', md, count=1)
    heads = [h.strip() for h in re.findall(r'^#{2,3} (.+)$', body, re.M)]

    ws = openpyxl.load_workbook(X_HELP).active
    cells = [str(ws.cell(r, 1).value or '') for r in range(1, ws.max_row + 1)]
    miss = [h for h in heads if h not in cells]
    check('12 派生ファイル', 'ヘルプExcelが最新のmdと同じ見出しを持つ', not miss, f'欠落 {miss[:5]}')

    html = open('20260805_従業員管理_05_ヘルプページ.html', encoding='utf-8').read()
    m = re.search(r'<script type="text/plain" id="src">([\s\S]*?)</script>', html)
    embedded = m.group(1) if m else ''
    # 見出しだけでなく本文まるごとを比べる。見出しが変わらないまま本文が古くなる
    # ことが実際に起きたため（2026-08-13）。
    check('12 派生ファイル', 'ヘルプHTMLの埋め込みがmdと一字一句同じ',
          bool(m) and embedded.strip() == body.strip(),
          f'md {len(body.strip())}文字 ／ HTML {len(embedded.strip())}文字'
          if m else 'src が無い')

    # ヘルプExcelも本文まるごとで比べる（見出し一致だけでは古さを検出できない）
    flat = '\n'.join(str(ws.cell(r, c).value or '')
                     for r in range(1, ws.max_row + 1)
                     for c in range(1, 4))
    plain = [re.sub(r'\*\*(.+?)\*\*', r'\1', l).strip()
             for l in body.split('\n')
             if l.strip() and not l.startswith('|') and not l.startswith('#')]
    miss = [l for l in plain if l.lstrip('> ') not in flat]
    check('12 派生ファイル', 'ヘルプExcelの本文がmdと一致', not miss,
          f'{len(plain) - len(miss)}/{len(plain)}行 欠落例 {miss[:2]}')

    # 撮影の点数が md の 📸 と一致するか（本文とExcelで数が食い違っていた）
    shots = re.findall(r'📸 \*\*(.+?)\*\*｜', body)
    wbh = openpyxl.load_workbook(X_HELP)
    ok = ('撮影一覧' in wbh.sheetnames
          and f'（{len(shots)}点）' in str(wbh['撮影一覧'].cell(1, 1).value or ''))
    check('12 派生ファイル', f'撮影一覧が {len(shots)}点', ok,
          str(wbh['撮影一覧'].cell(1, 1).value or '')[:40] if '撮影一覧' in wbh.sheetnames else 'シートが無い')

    # ---- 検証データセットExcel（data/ と検証プランの写し）----
    X_DATA = '20260805_従業員管理_04_検証データセット.xlsx'
    wbd = openpyxl.load_workbook(X_DATA)
    files = sorted(f for f in os.listdir('data') if f.endswith(('.tsv', '.csv')))
    miss = [f for f in files if os.path.splitext(f)[0] not in wbd.sheetnames]
    check('12 派生ファイル', f'検証データセットExcelに data/ の{len(files)}件が揃う',
          not miss, f'シートが無い: {miss}')
    bad = []
    for f in files:
        nm = os.path.splitext(f)[0]
        if nm not in wbd.sheetnames:
            continue
        d = '\t' if f.endswith('.tsv') else ','
        with open(os.path.join('data', f), encoding='utf-8-sig', newline='') as fh:
            rows = list(csv.reader(fh, delimiter=d))
        wsd = wbd[nm]
        hdr = [str(c.value) for c in wsd[3] if c.value is not None]
        if hdr != rows[0] or wsd.max_row - 3 != len(rows) - 1:
            bad.append(nm)
    check('12 派生ファイル', '検証データセットExcelの列・行数が data/ と一致',
          not bad, f'ずれ: {bad}')
    prep = [r['検証ID'] for r in T(F_PLAN) if r['段階'] == '準備']
    got = [str(wbd['01_投入手順'].cell(r, 1).value or '')
           for r in range(4, wbd['01_投入手順'].max_row + 1)]
    check('12 派生ファイル', f'投入手順が検証プランの準備{len(prep)}行と一致',
          [g for g in got if g] == prep, f'Excel {[g for g in got if g][:3]} / TSV {prep[:3]}')

    # ---- モックの表のマークアップ ----
    # thead が入れ子だと最初の行が空になり、DataTables が
    # 「Incorrect column count」で初期化に失敗する。初期化が止まると同じ
    # $(document).ready 内の後続処理（給与計算モードの切替など）も動かない。
    nest, mismatch = [], []
    for f in sorted(glob.glob('mock/*.html')):
        h = open(f, encoding='utf-8').read()
        for m in re.finditer(r'<thead>[\s\S]*?</thead>', h):
            if '<thead>' in m.group(0)[7:]:
                nest.append(os.path.basename(f))
        for m in re.finditer(r'<table[^>]*id="([^"]+)"[\s\S]*?</table>', h):
            seg = m.group(0)
            if '</thead>' not in seg:
                continue
            nth = len(re.findall(r'<th[\s>]', seg[:seg.index('</thead>')]))
            ntd = {len(re.findall(r'<td[\s>]', r))
                   for r in re.findall(r'<tr[^>]*>[\s\S]*?</tr>', seg[seg.index('</thead>'):])
                   if '<td' in r}
            if nth and ntd and {nth} != ntd:
                mismatch.append(f'{os.path.basename(f)}#{m.group(1)} th={nth} td={sorted(ntd)}')
    check('7 ヘルプ', 'モックの thead が入れ子になっていない', not nest, sorted(set(nest)))
    check('7 ヘルプ', 'モックの表で見出しとデータの列数が一致', not mismatch, mismatch[:4])

    # 給与実績の3画面で、概算モードに残す列がそろっている（判断22）
    KEEP = {'名前', '従業員ページ', '従業員設定', '従業員コード', '年月', '日時', '雇用区分',
            '所属店舗', '給与単位', '単位給与額', '想定労働日数', '想定労働時間',
            '想定勤務時間', 'みなし労働時間', 'みなし残業時間', '想定給与'}
    bad = []
    for f in ['mock/payroll_reports.html', 'mock/employee_detail.html',
              'mock/employee_payroll_detail.html']:
        h = open(f, encoding='utf-8').read()
        for tab in ['monthly_tab', 'daily_tab']:
            j = h.find(f'id="{tab}"')
            if j < 0:
                continue
            seg = h[j:h.index('</thead>', j)]
            for tag, name in re.findall(r'(<th(?![a-z])[^>]*>)(?:<[^>]+>)*([^<]+)', seg):
                nm = name.strip()
                if not nm:
                    continue
                if (nm not in KEEP) != ('actual' in tag):
                    bad.append(f'{os.path.basename(f)}/{tab}/{nm}')
    check('7 ヘルプ', '概算モードで残す列が3画面6タブで同じ（判断22）', not bad, bad[:6])

    # 行ごとに、見出しとセルがそろっているか。列を足したときに値だけ別の位置に
    # 入る事故が実際に起きた（金額の列に「時間」が出た）。
    def money(x):
        m = re.search(r'([+\-]?[\d,]+)円', re.sub(r'<[^>]+>', '', x))
        return int(m.group(1).replace(',', '')) if m else None

    bad = []
    for f in ['mock/payroll_reports.html', 'mock/employee_detail.html',
              'mock/employee_payroll_detail.html']:
        h = open(f, encoding='utf-8').read()
        for tab in ['monthly_tab', 'daily_tab']:
            j = h.find(f'id="{tab}"')
            if j < 0:
                continue
            seg = h[j:h.index('</table>', j)]
            hd = seg.index('</thead>')
            hs = [('actual' in m.group(1), re.sub(r'<[^>]+>', '', m.group(2)).strip())
                  for m in re.finditer(r'(?s)(<th(?![a-z])[^>]*>)(.*?)</th>', seg[:hd])]
            names = [n for _, n in hs]
            for rn, r in enumerate(re.findall(r'<tr[^>]*>[\s\S]*?</tr>', seg[hd:]), 1):
                cells = re.findall(r'<td[^>]*>[\s\S]*?</td>', r)
                tag = f'{os.path.basename(f)}/{tab}/{rn}行'
                if len(cells) != len(hs):
                    bad.append(f'{tag} td={len(cells)} th={len(hs)}')
                    continue
                for (a, nm), c in zip(hs, cells):
                    v = re.sub(r'<[^>]+>', '', c)
                    if a != ('actual' in c):
                        bad.append(f'{tag} 「{nm}」の印')
                    if ('給与' in nm or '人件費' in nm or '残業代' in nm) and ('時間' in v or '日' in v):
                        bad.append(f'{tag} 金額の列「{nm}」に {v.strip()[:8]}')
                    if ('時間' in nm or '日数' in nm) and '円' in v:
                        bad.append(f'{tag} 時間の列「{nm}」に {v.strip()[:8]}')
                if {'想定給与', '給与', '給与実績差分'} <= set(names):
                    e_, p_, d_ = (money(cells[names.index(k)])
                                  for k in ['想定給与', '給与', '給与実績差分'])
                    if None not in (e_, p_, d_) and p_ - e_ != d_:
                        bad.append(f'{tag} 差分{d_} ≠ {p_}-{e_}')
    check('7 ヘルプ', '給与実績の表で行と列がそろっている', not bad, bad[:6])

    # トップページ（index.html）の件数と画面一覧
    idx = open('index.html', encoding='utf-8').read()
    ac = T(F_AC)
    tp = T(F_PLAN)
    lv = collections.Counter(a['判定レベル'] for a in ac)
    n_l0, n_l1 = lv['L0 リリース不可'], lv['L1 MVP必須']
    prep = sum(1 for r in tp if r['段階'] == '準備')
    jl = open('JUDGMENT_LOG.md', encoding='utf-8').read()
    n_data = len(os.listdir('data'))
    for want, label in [(f'L0（{n_l0}件）', 'L0'), (f'L1（{n_l1}件）', 'L1'),
                        (f'L2（{lv["L2 次段階"]}件）', 'L2'),
                        (f'準備{prep}＋{len(tp) - prep}件', '検証プランのタグ'),
                        (f'未確定{jl.count("❌ 未確定")}件', '未確定の判断'),
                        (f'検証用・{n_data}ファイル', 'データセット')]:
        check('8 件数表記', f'index.htmlが {want}', want in idx, label)
    mocks = {os.path.basename(x) for x in glob.glob('mock/*.html')}
    linked = set(re.findall(r'href="mock/([^"?#]+\.html)"', idx))
    part = {'sidebar_company.html', 'header_company.html', 'index.html'}
    check('8 件数表記', 'index.htmlが全モック画面へリンクしている',
          not (mocks - linked - part), sorted(mocks - linked - part))

    # ビューアが参照するファイルの実在
    bad = []
    for f in glob.glob('*.html'):
        h = open(f, encoding='utf-8').read()
        for ref in set(re.findall(r"var f = '([^']+)'", h)) | set(re.findall(r"fetch\(encodeURI\('([^']+)'\)", h)):
            if not os.path.exists(ref):
                bad.append(f'{f} → {ref}')
    check('12 派生ファイル', 'ビューアHTMLの参照先が実在', not bad, bad[:5])


# ============================================================
def main():
    for fn in [c1_files, c2_acceptance, c3_plan, c4_links, c5_order, c6_data, c7_help, c8_counts, c9_layout,
               c10_ids, c11_wording, c12_derived]:
        try:
            fn()
        except Exception as e:
            check(fn.__name__, '検査自体が失敗', False, f'{type(e).__name__}: {e}')

    cat = None
    for c, name, ok, detail in results:
        if c != cat:
            print(f'\n■ {c}')
            cat = c
        mark = '✅' if ok else '❌'
        print(f'  {mark} {name}' + (f'  … {detail}' if (detail and not ok) else (f'  … {detail}' if detail else '')))

    ng = [r for r in results if not r[2]]
    print('\n' + '=' * 64)
    print(f'  検査 {len(results)}件 ／ OK {len(results)-len(ng)}件 ／ NG {len(ng)}件')
    if ng:
        print('\n  要対応:')
        for c, name, _, detail in ng:
            print(f'    ・[{c}] {name}  {detail}')
    else:
        print('  すべての資料が整合しています。')
    print('=' * 64)
    return 1 if ng else 0


if __name__ == '__main__':
    sys.exit(main())

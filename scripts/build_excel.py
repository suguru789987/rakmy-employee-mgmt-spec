#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配布用Excelを TSV / Markdown から生成する。

使い方:  python3 scripts/build_excel.py
生成するもの:
  20260805_従業員管理_00_定義表.xlsx      … 言葉の定義（受入条件表TSVから件数を集計）
  20260805_従業員管理_01_確認表.xlsx      … 受入条件表＋記入欄（受入条件表TSV・確認表TSVから）
  20260805_従業員管理_02_検証プラン.xlsx  … 検証プラン（検証プランTSVから）

TSVを直したら必ずこれを流す。手でExcelを編集すると次の生成で消える。
"""
import csv, math, os, re, sys, collections

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)

F_AC, F_CHK, F_PLAN = ('20260803_02_従業員管理_受入条件表.tsv',
                       '20260803_02b_従業員管理_受入条件_確認表.tsv',
                       '20260803_03_従業員管理_検証プラン.tsv')

HDR = PatternFill('solid', fgColor='F1F3F5')
YEL = PatternFill('solid', fgColor='FFF9DB')
RED = PatternFill('solid', fgColor='FFEBE9')
GRN = PatternFill('solid', fgColor='DAFBE1')
BLU = PatternFill('solid', fgColor='DDF4FF')
GRY = PatternFill('solid', fgColor='F6F8FA')
SHOT = PatternFill('solid', fgColor='FFF3BF')
SCR = PatternFill('solid', fgColor='EAF4FF')
LNK = PatternFill('solid', fgColor='EDE7FF')
DAT = PatternFill('solid', fgColor='E8F5E9')
PREP = PatternFill('solid', fgColor='FFFBE6')
REV = PatternFill('solid', fgColor='F3F0FF')
CAL = PatternFill('solid', fgColor='EAFBEA')
THIN = Side(style='thin', color='D0D7DE')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE = Font(bold=True, size=13)
NOTE = Font(size=10, color='57606A')
BLK = Font(bold=True, size=11)

MS = collections.OrderedDict([
    ('実装M1 データ基盤', ('M1 過去が動かない構造ができる', '設定が適用日付つきで保存され、履歴が積まれる。同じ従業員コードが二重に登録されない')),
    ('実装M2 計算エンジン', ('M2 金額が算式どおりに出る', '確定した算式で人件費が計算され、店舗別の合計が全社と一致する')),
    ('実装M3 締め', ('M3 月次を確定して外部へ渡せる', '締めた月は動かず、解除すれば再計算される。締めは会社スコープのみ')),
    ('実装M4 画面・入出力', ('M4 顧客とCSが自分で運用できる', 'CSVで取り込め、必要な項目だけ出力でき、数字がおかしいときに原因を追える')),
    ('実装M5 移行', ('M5 既存顧客を壊さない', '移行しても既存データと過去月の人件費が変わらない')),
    ('（次リリース）', ('M6 次段階（今回は実装しない）', '無くても業務は回る。実測を記録し次の優先順位を決める')),
])


def T(p):
    with open(p, encoding='utf-8-sig') as f:
        return list(csv.DictReader(f, delimiter='\t'))


def lines(txt, width):
    if not txt:
        return 1
    n = 0
    for seg in str(txt).split('\n'):
        w = sum(2 if ord(c) > 0x7f else 1 for c in seg)
        n += max(1, math.ceil(w / max(4, width - 2)))
    return n


# ============================================================ 00 定義表
def build_def():
    d = T(F_AC)
    tp = T(F_PLAN)
    lv = collections.Counter(a['判定レベル'] for a in d)
    ms = collections.Counter(a['実装マイルストーン'] for a in d)
    cm = collections.Counter(a['確認方法'] for a in d)
    pt, vm = collections.OrderedDict(), collections.OrderedDict()
    for a in d:
        pt.setdefault(a['守る設計'].split('：')[0], []).append(a['条件ID'])
        vm.setdefault(a['価値MVP'], []).append(a['条件ID'])
    n_data = len([f for f in os.listdir('data')])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '定義'
    W = [26, 10, 92, 18]
    for j, w in enumerate(W, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.cell(1, 1, '従業員管理 受入条件表 ─ 定義').font = TITLE
    ws.cell(2, 1, '言葉の定義だけを載せています。実際の判定は 20260805_従業員管理_01_確認表.xlsx で行います。').font = NOTE
    r = 4

    def blk(r, title, rows, note=None):
        c = ws.cell(r, 1, title); c.font = BLK; c.fill = HDR
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(W))
        ws.row_dimensions[r].height = 20; r += 1
        if note:
            c = ws.cell(r, 1, note); c.font = NOTE
            c.alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(W))
            ws.row_dimensions[r].height = 15 * lines(note, sum(W)) + 4; r += 1
        for i, row in enumerate(rows):
            need = 1
            for j, v in enumerate(row, 1):
                cc = ws.cell(r + i, j, v); cc.border = BOX
                cc.font = Font(bold=(i == 0), size=10)
                cc.alignment = Alignment(wrap_text=True, vertical='top')
                if i == 0:
                    cc.fill = HDR
                need = max(need, lines(v, W[j - 1]))
            ws.row_dimensions[r + i].height = 15 * need + 4
        return r + len(rows) + 1

    r = blk(r, '■ 1. 4つの文書の役割分担', [
        ['層', '文書', '何を決めるか', '判断の主体'],
        ['1. 意図', '仕様書 20260803_01_..._PdM版.md', 'なぜそう作るか。11フローのケース別遷移。判断の記録', 'PdM'],
        ['2. 合格ライン', '受入条件表 20260805_01_確認表.xlsx', '意図を満たしたと言える定量条件。実装レベルとマイルストーン', 'PdM＋エンジニア'],
        ['3. 挙動', '検証プラン 20260805_02_検証プラン.xlsx', 'その条件を実際の画面・データでどう確かめるか', '検証実施者'],
        ['4. 前提データ', f'検証データセット data/01〜{n_data}（{n_data}ファイル）', '検証に必要な入力データ。算式で再現できる値', '検証実施者'],
    ], '2と3は独立して使えます。受入条件表は合格ラインと確認する画面だけで判定でき、検証プランは操作と期待値だけで実施できます。相互参照は「対応検証ID」「対応受入条件」で行います。')
    r = blk(r, '■ 2. 実装レベルの定義', [
        ['実装レベル', '件数', '意味'],
        ['L0 リリース不可', lv['L0 リリース不可'], '1件でも落ちるとリリースしない。過去が動く／締めても動く／店舗別が全社と合わない、のいずれかが起きている状態'],
        ['L1 MVP必須', lv['L1 MVP必須'], '業務が回る最低ライン。「実装できていないと起きること」が【不可】なら落とせない、【条件付きGo可】なら手作業で代替できる'],
        ['L2 次段階', lv['L2 次段階'], '今回は実装しない。実測を記録し次の優先順位を決める'],
    ])
    r = blk(r, '■ 3. マイルストーンの定義（実装順）',
            [['マイルストーン', '条件数', '完了条件']] +
            [[v[0], ms[k], v[1]] for k, v in MS.items()],
            'リソースが足りない場合の確保順は M1 → M5 → M2 → M3 → M4。M1を飛ばして画面から作ると、その期間のデータは適用日付と履歴を持たず、あとから正しい過去を作れません。')
    r = blk(r, '■ 4. 設計ポイントの定義', [['設計ポイント', '条件数', '含まれる条件ID']] +
            [[k, len(v), ' ／ '.join(v)] for k, v in pt.items()])
    r = blk(r, '■ 5. 価値MVPの定義（これが成立して初めて価値が出る）',
            [['価値', '条件数', '成立した状態と、欠けたときに起きること']] +
            [[k, len(v), {
                'V1 数字が動かない': '過去の月も確定した月もあとから金額が変わらない。欠けると数字が信用されずExcelに戻る',
                'V2 数字が実態を表す': '金額が算式どおりで、店舗別が実際の働き方と一致する。欠けると店舗別PLが使えない',
                'V3 自分で運用できる': '導入・説明・外部提出を開発に依頼せず回せる。欠けると運用が開発待ちになる',
            }.get(k, '')] for k, v in vm.items()])
    r = blk(r, '■ 6. 確認方法の定義', [
        ['確認方法', '件数', '意味'],
        ['操作で確認', cm['操作で確認'], '画面を操作すれば判定できる'],
        ['操作＋実装レビュー', cm['操作＋実装レビュー'], '画面を操作しても実装の是非を判別できない。実装物（スキーマ・API・コード）を提示してもらって確認する（検証プラン R-01〜R-06）'],
    ])
    r = blk(r, '■ 7. 判断の状態（2026-08-13 時点）', [
        ['区分', '件数', '内容'],
        ['確定', 8, '判断1 所定労働日割 ／ 判断5 ヘルプ人件費は振替 ／ 判断6 交通費は独立費目（2026-08-03）　判断7 深夜割増1.5 ／ 判断8 みなし超過1.25・未達は減額なし ／ 判断9 端数は所属店舗に寄せる ／ 判断10 設定編集の入口はA（両方残す・更新処理を1箇所に集約）（2026-08-07）　判断11 設定変更履歴は直近3件・降順・ステータス2列目・変更点を強調（2026-08-13）'],
        ['未確定', 3, '判断2 翌月差額調整の費目 ／ 判断3・4 画面名の統一（いずれもリリースまでに確定させる）'],
    ])
    ws.freeze_panes = 'A3'
    ws.sheet_view.showGridLines = False
    wb.save('20260805_従業員管理_00_定義表.xlsx')
    return len(d)


# ============================================================ 01 確認表
def build_check():
    d = T(F_AC)
    c = {x['条件ID']: x for x in T(F_CHK)}
    COLS = ['マイルストーン', 'マイルストーンの完了条件', '設計ポイント', '条件ID', '実装レベル', '何を実装するか',
            '合格ライン（この数値を満たせば実装完了）', '実装できていないと起きること', '実装箇所', '確認する画面',
            '対応検証ID', 'ヘルプページ該当箇所', '仕様書の根拠', 'エビデンス', '合否', '実測値', '判定者', '判定日']
    W = [22, 30, 22, 9, 12, 38, 44, 40, 34, 22, 16, 30, 32, 14, 9, 20, 11, 11]
    HR = 12
    top, bot = HR + 1, HR + len(d)
    OK = get_column_letter(COLS.index('合否') + 1)

    # 条件IDごとのヘルプ該当箇所（検証プランから集約）
    help_map = collections.defaultdict(list)
    for r in T(F_PLAN):
        h = r['ヘルプページ該当箇所']
        if not h or h.startswith('（'):
            continue
        for ac in re.findall(r'AC-\d+', r['対応受入条件']):
            if h not in help_map[ac]:
                help_map[ac].append(h)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '受入条件表'
    ws.cell(1, 1, '従業員管理 受入条件表 ─ どこまで実装できれば合格か').font = TITLE
    ws.cell(2, 1, '「合否」に OK / NG を入れると下の到達状況が自動更新されます').font = NOTE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))

    for j, v in enumerate(['マイルストーン', '完了条件', '', '', '条件数', 'OK', 'NG', '', '', '状態'], 1):
        cc = ws.cell(4, j, v); cc.font = Font(bold=True, size=10); cc.fill = HDR; cc.border = BOX
    for i, (k, v) in enumerate(MS.items()):
        r = 5 + i
        ws.cell(r, 1, v[0]).font = Font(size=10)
        ws.cell(r, 2, v[1]).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(r, 5, f'=COUNTIF($A${top}:$A${bot},$A{r})')
        ws.cell(r, 6, f'=COUNTIFS($A${top}:$A${bot},$A{r},${OK}${top}:${OK}${bot},"OK")')
        ws.cell(r, 7, f'=COUNTIFS($A${top}:$A${bot},$A{r},${OK}${top}:${OK}${bot},"NG")')
        ws.cell(r, 10, f'=IF($G{r}>0,"未達（NG "&$G{r}&"件）",IF($F{r}=$E{r},"完了",'
                       f'"進行中（残 "&($E{r}-$F{r})&"件）"))')
        for j in range(1, 11):
            ws.cell(r, j).border = BOX
        ws.row_dimensions[r].height = 15 * lines(v[1], W[1]) + 4
    ws.cell(11, 1, '総合判定').font = BLK
    ws.cell(11, 2, f'=IF(COUNTIFS($E${top}:$E${bot},"L0*",${OK}${top}:${OK}${bot},"NG")>0,"✗ リリース不可（L0にNG）",'
                   f'IF(COUNTIFS($E${top}:$E${bot},"L1*",${OK}${top}:${OK}${bot},"NG")>0,"△ 設計MVP未達（L1にNG）",'
                   f'IF(COUNTIFS($E${top}:$E${bot},"L0*",${OK}${top}:${OK}${bot},"OK")+COUNTIFS($E${top}:$E${bot},"L1*",${OK}${top}:${OK}${bot},"OK")'
                   f'<COUNTIF($E${top}:$E${bot},"L0*")+COUNTIF($E${top}:$E${bot},"L1*"),"進行中（L0・L1に未判定あり）","◎ リリース可（L0・L1すべてOK）")))').font = BLK
    ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=10)

    for j, v in enumerate(COLS, 1):
        x = ws.cell(HR, j, v); x.font = Font(bold=True, size=10); x.fill = HDR; x.border = BOX
        x.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.row_dimensions[HR].height = 34

    for i, a in enumerate(d):
        r = HR + 1 + i
        m = MS.get(a['実装マイルストーン'], (a['実装マイルストーン'], ''))
        cid = a['条件ID']
        vid = c.get(cid, {}).get('検証ID', '')
        ev = (vid.split()[0] + '.png') if vid and vid != '-' else '-'
        vals = [m[0], m[1], a['守る設計'], cid, a['判定レベル'], a['満たすべき状態'],
                a['合格基準（定量）'], a['不合格のときリリースは'], a['実装の目安（どこまでできていれば合格か）'],
                a['対象画面'], vid, ' ／ '.join(help_map.get(cid, [])), a.get('仕様書の根拠', ''), ev, '', '', '', '']
        need = 1
        for j, v in enumerate(vals, 1):
            x = ws.cell(r, j, v); x.border = BOX; x.font = Font(size=10)
            wrap = j not in (4, 5, 15, 17, 18)
            x.alignment = Alignment(wrap_text=wrap, vertical='top')
            if j >= 15:
                x.fill = YEL
            if wrap:
                need = max(need, lines(v, W[j - 1]))
        ws.row_dimensions[r].height = 15 * need + 4
        lvcell = ws.cell(r, 5)
        if a['判定レベル'].startswith('L0'):
            lvcell.font = Font(bold=True, size=10, color='CF222E'); lvcell.fill = RED
        elif a['判定レベル'].startswith('L2'):
            lvcell.fill = GRY
        else:
            lvcell.fill = BLU

    for j, w in enumerate(W, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(top, 4)
    ws.auto_filter.ref = f'A{HR}:{get_column_letter(len(COLS))}{bot}'
    dv = DataValidation(type='list', formula1='"OK,NG,対象外"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f'{OK}{top}:{OK}{bot}')
    ws.conditional_formatting.add(f'{OK}{top}:{OK}{bot}', CellIsRule(operator='equal', formula=['"OK"'], fill=GRN))
    ws.conditional_formatting.add(f'{OK}{top}:{OK}{bot}', CellIsRule(operator='equal', formula=['"NG"'], fill=RED))
    ws.conditional_formatting.add(f'J5:J{4+len(MS)}', FormulaRule(formula=['LEFT($J5,2)="完了"'], fill=GRN))
    ws.conditional_formatting.add(f'J5:J{4+len(MS)}', FormulaRule(formula=['LEFT($J5,2)="未達"'], fill=RED))
    ws.sheet_view.showGridLines = False
    wb.save('20260805_従業員管理_01_確認表.xlsx')
    return len(d), len(COLS)


# ============================================================ 02 検証プラン
def build_plan():
    tp = T(F_PLAN)
    prep = [r for r in tp if r['段階'] == '準備']
    COLS = ['段階', '検証ID', '検証内容', '操作・前提条件／入力値', '使用する検証データ', '期待値',
            'スクショファイル名', 'スクショ（撮るもの）', 'スクショ貼付欄', '画面', '確認ポイント',
            'ヘルプページ該当箇所', '対応受入条件', '実際の値', '判定', '所見・気づき', '実施日', '実施者', '計算根拠']
    W = [12, 10, 22, 58, 28, 58, 15, 44, 38, 50, 24, 30, 18, 20, 10, 28, 12, 12, 42]
    WRAP = {3, 4, 5, 6, 8, 10, 11, 12, 13, 16, 19}
    REC = {9, 14, 15, 16, 17, 18}

    def op_text(r):
        pre = r['初期設定（前提データ）'].strip()
        return f"【前提】{pre}\n【操作】{r['操作・前提条件']}" if pre and pre != '-' else r['操作・前提条件']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '検証プラン'
    NOTE_TXT = (f'確かめる手順（準備{len(prep)}行＋検証{len(tp)-len(prep)}ケース）。上から順に実施する。'
                '段階でフィルタすると 一覧／登録／詳細／履歴／実績／締め／取込／出力／取込データ／横断／計算／実装レビュー に絞り込める。'
                '\n実施順を変えてはいけない箇所：①T-A08（勤怠取込）→T-S03（締め）　②T-S03→T-S04・T-D07→T-S05（この3件は締め済みでしか確認できない）　'
                '③C-000（初期化）は計算の先頭　④C-008は計算の最後（2026-05のEMP-001を352,286円に変えるためC-013・C-017が合わなくなる）　'
                '⑤T-N02の前にdata/02を再投入　⑥削除はEMP-012（T-L10）とEMP-004（T-D08）だけ'
                '\n列の使い方：使用する検証データ＝この検証で投入・使用するdataファイル／期待値＝先頭の【確認場所】でどの画面を見て合否を出すかを示し、続けて合格の基準を書いている／'
                'スクショ（撮るもの）＝何が写っていれば証拠になるか／スクショ貼付欄＝撮った画像をこの列に貼る／'
                '対応受入条件＝IDで受入条件表を引くと合格基準（定量）がある。検証後はIDごとに確認表へ集約し、紐づく検証がすべてOKで初めて合格／'
                'ヘルプページ該当箇所＝顧客への約束。NGのとき実装を直すのかヘルプを直すのかを決める'
                '\n色：緑＝使う検証データ　青＝画面　橙＝スクショの指示　紫＝ヘルプ・受入条件との対応　黄＝当日の記入欄。行の地色は 薄黄＝準備　薄緑＝計算　薄紫＝実装レビュー')
    ws.cell(1, 1, NOTE_TXT).font = NOTE
    ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    ws.row_dimensions[1].height = 15 * lines(NOTE_TXT, sum(W)) + 8
    for j, v in enumerate(COLS, 1):
        c = ws.cell(2, j, v); c.font = Font(bold=True, size=10); c.fill = HDR; c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.row_dimensions[2].height = 32
    for i, r0 in enumerate(tp):
        r = 3 + i
        vals = [r0['段階'], r0['検証ID'], r0['検証内容'], op_text(r0), r0['使用する検証データ'], r0['期待値'],
                r0['スクショファイル名'], r0['スクショ（撮るもの）'], '', r0['画面'], r0['確認ポイント'],
                r0['ヘルプページ該当箇所'], r0['対応受入条件'], '', '', '', '', '', r0['計算根拠']]
        bg = {'準備': PREP, '計算': CAL, '実装レビュー': REV}.get(r0['段階'])
        need = 1
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.border = BOX; c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=(j in WRAP), vertical='top')
            if j in REC:
                c.fill = YEL
            elif bg:
                c.fill = bg
            elif j == 5:
                c.fill = DAT
            elif j in (7, 8):
                c.fill = SHOT
            elif j == 10:
                c.fill = SCR
            elif j in (12, 13):
                c.fill = LNK
            if j in WRAP:
                need = max(need, lines(v, W[j - 1]))
        ws.row_dimensions[r].height = max(60, 15 * need + 4)
    for j, w in enumerate(W, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'C3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(COLS))}{2+len(tp)}'
    dv = DataValidation(type='list', formula1='"OK,NG,対象外"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f'O3:O{2+len(tp)}')
    ws.sheet_view.showGridLines = False
    wb.save('20260805_従業員管理_02_検証プラン.xlsx')
    return len(tp), len(COLS)


if __name__ == '__main__':
    n = build_def()
    print(f'  OK 00_定義表.xlsx（受入条件{n}件から集計）')
    n, c = build_check()
    print(f'  OK 01_確認表.xlsx（{n}条件・{c}列）')
    n, c = build_plan()
    print(f'  OK 02_検証プラン.xlsx（{n}行・{c}列）')

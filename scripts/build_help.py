#!/usr/bin/env python3
# 20260805_従業員管理_05_ヘルプページ.html / .xlsx を マニュアル版.md から作り直す。
#
# なぜ自動生成にするか:
#   この2つは md を手で写したものだった。写しなので md を直しても追随せず、
#   2026-08-13 時点で 8/7 の内容のまま止まっていた（旧画面名「人件費実績」が残り、
#   想定労働日数・日別の概算モードが入っていない）。手で直すのをやめ、毎回作り直す。
#
# 使い方:  python3 scripts/build_help.py

import os
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MD = os.path.join(ROOT, '20260805_ヘルプページ_従業員管理_マニュアル版.md')
HTML = os.path.join(ROOT, '20260805_従業員管理_05_ヘルプページ.html')
XLSX = os.path.join(ROOT, '20260805_従業員管理_05_ヘルプページ.xlsx')
LOG = os.path.join(ROOT, 'JUDGMENT_LOG.md')

# 掲載前に確認が要る箇所と、それが待っている判断。状態は JUDGMENT_LOG から拾う。
BLOCKERS = [
    ('月の途中で給与が変わる場合の計算例', 'J-01', '所定労働日割'),
    ('ヘルプ勤務の人件費の計上先', 'J-05', '振替（応援先に加算し所属店舗から差引）'),
    ('交通費が人件費に含まれるか', 'J-06', '独立費目（通勤は所属店舗・応援先ぶんは応援先）'),
    ('深夜残業代の金額', 'J-07', '割増率'),
    ('みなし残業を超えた／届かなかったときの扱い', 'J-08', '超過は×1.25・未達は減額なし'),
    ('給与実績一覧のモーダルから直したときの適用日付', 'J-10', '設定編集の入口を2つ残すか'),
    ('日次の想定給与と概算の日別', 'J-20', '想定労働日数の出所'),
    ('日別の概算モードの説明', 'J-21', '概算の日別は営業日でならす'),
    ('翌月で差額を調整する条件', 'J-02', '対象費目'),
    ('「設定インポート」「設定取込データ」の画面名', 'J-03', '名称の統一'),
    ('画面間のリンク名', 'J-04', 'J-03と同時に決定'),
]

# 実装が終わらないと書けない箇所（判断ではなく実装の確認）
IMPL = [
    ('予約した変更の取り消し手順', '実装されているか、手順は何か'),
    ('従業員を削除したときの過去実績の扱い', '実績が残るか、消えるか'),
    ('締め解除の記録（誰が・いつ）', '記録が残るか（検証プラン T-S05 の所見を参照）'),
    ('締め済み期間への適用時の警告表示', '警告が出るか（モック未実装。T-D07）'),
    ('締め実行時の事前チェック', '設定が未入力の従業員がいるときに警告が出るか（T-S03の所見・次段階）'),
    ('CSV取り込みのエラー表示', '行単位でエラーになるか、全体が拒否されるか（T-I05）'),
]

TITLE = Font(bold=True, size=12)
HEAD = Font(bold=True, size=10)
H2 = Font(bold=True, size=11, color='1F883D')
H3 = Font(bold=True, size=10, color='0969DA')
BODY = Font(size=10)
FILL = PatternFill('solid', fgColor='F1F3F5')
LINE = Border(bottom=Side(style='thin', color='CED4DA'))
WRAP = Alignment(vertical='top', wrap_text=True)
TOP = Alignment(vertical='top')


def judgment_status():
    """JUDGMENT_LOG の一覧表から J-xx の確定状態を読む。"""
    txt = open(LOG, encoding='utf-8').read()
    out = {}
    for m in re.finditer(r'^\| \*\*(J-\d+)\*\* \|(.*)$', txt, re.M):
        row = m.group(2)
        fixed = '✅' in row
        date = re.search(r'(\d{4}-\d{2}-\d{2})', row)
        out[m.group(1)] = (f'✅ {date.group(1)} 確定・本文へ反映済み' if fixed
                           else '❌ 未確定（この箇所は書けません）')
    return out


def md_body():
    md = open(MD, encoding='utf-8').read()
    return re.sub(r'<!--[\s\S]*?-->\n*', '', md, count=1).strip()


def build_html(body):
    html = open(HTML, encoding='utf-8').read()
    start = html.index('<script type="text/plain" id="src">') + len('<script type="text/plain" id="src">')
    end = html.index('</script>', start)
    open(HTML, 'w', encoding='utf-8').write(html[:start] + '\n' + body + '\n' + html[end:])
    print(f'  ✅ {os.path.basename(HTML)}  本文 {len(body)}文字を埋め込み')


def build_xlsx(body):
    wb = openpyxl.Workbook()

    # ---- 本文 ----
    ws = wb.active
    ws.title = 'ヘルプページ本文'
    n = 1
    for raw in body.split('\n'):
        line = raw.rstrip()
        if not line.strip():
            continue
        if re.fullmatch(r'\|[\s|:-]+\|', line):      # 表の区切り行
            continue
        if line.startswith('|'):
            cells = [c.strip() for c in line.strip('|').split('|')]
            for i, v in enumerate(cells, 1):
                c = ws.cell(n, i, re.sub(r'\*\*(.+?)\*\*', r'\1', v))
                c.font, c.alignment = BODY, WRAP
        else:
            m = re.match(r'^(#{1,3}) (.+)$', line)
            text = re.sub(r'\*\*(.+?)\*\*', r'\1', m.group(2) if m else line.lstrip('> '))
            c = ws.cell(n, 1, text)
            c.font = {1: TITLE, 2: H2, 3: H3}[len(m.group(1))] if m else BODY
            c.alignment = WRAP
        n += 1
    for col, w in zip('ABC', [40, 52, 38]):
        ws.column_dimensions[col].width = w

    # ---- 撮影一覧（本文の 📸 から作る）----
    shots = re.findall(r'📸 \*\*(.+?)\*\*｜(.+?)(?:\n|$)', body)
    ws = wb.create_sheet('撮影一覧')
    ws.cell(1, 1, f'ヘルプページ 撮影が必要なスクリーンショット（{len(shots)}点）').font = TITLE
    for i, h in enumerate(['#', '画面', '撮影の指示'], 1):
        c = ws.cell(3, i, h)
        c.font, c.fill, c.border, c.alignment = HEAD, FILL, LINE, TOP
    for i, (name, how) in enumerate(shots, 1):
        for j, v in enumerate([i, name, how], 1):
            c = ws.cell(3 + i, j, v)
            c.font, c.alignment = BODY, WRAP
        ws.row_dimensions[3 + i].height = 14 * (len(how) // 44 + 1) + 4
    for col, w in zip('ABC', [6, 34, 74]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A4'

    # ---- 掲載前チェック（社内用）----
    st = judgment_status()
    ws = wb.create_sheet('掲載前チェック（社内用）')
    ws.cell(1, 1, 'ヘルプページ 掲載前チェック（社内用・顧客には出しません）').font = TITLE
    n = 2
    for i, h in enumerate(['#', '箇所', '待っている判断', '状態'], 1):
        c = ws.cell(n, i, h)
        c.font, c.fill, c.border, c.alignment = HEAD, FILL, LINE, TOP
    for i, (where, jid, what) in enumerate(BLOCKERS, 1):
        for j, v in enumerate([i, where, f'{jid.replace("J-0", "判断").replace("J-", "判断")}＝{what}',
                               st.get(jid, '（判断記録に無い）')], 1):
            c = ws.cell(n + i, j, v)
            c.font, c.alignment = BODY, WRAP
    n += len(BLOCKERS) + 2
    ws.cell(n, 1, '実装確認が必要な箇所').font = TITLE
    n += 1
    for i, h in enumerate(['#', '箇所', '確認すること'], 1):
        c = ws.cell(n, i, h)
        c.font, c.fill, c.border, c.alignment = HEAD, FILL, LINE, TOP
    for i, (where, what) in enumerate(IMPL, 1):
        for j, v in enumerate([i, where, what], 1):
            c = ws.cell(n + i, j, v)
            c.font, c.alignment = BODY, WRAP
    n += len(IMPL) + 2
    ws.cell(n, 1, '撮影').font = TITLE
    ws.cell(n + 1, 1, f'撮影が必要なスクリーンショットは「撮影一覧」シート（{len(shots)}点）を見てください。'
                      '本文の 📸 から自動で作っているため、本文と必ず一致します。').alignment = WRAP
    ws.cell(n + 2, 1, '撮影は実装完了後の画面で行ってください。'
                      'モックの画面をそのまま掲載すると、実装との差異が顧客に見えます。').alignment = WRAP
    n += 4
    ws.cell(n, 1, '掲載の可否').font = TITLE
    open_j = [b for b in BLOCKERS if not st.get(b[1], '').startswith('✅')]
    ws.cell(n + 1, 1, f'未確定の判断が {len(open_j)}件 残っています：'
                      + '／'.join(f'{b[1]} {b[0]}' for b in open_j) if open_j
                      else '判断はすべて確定しています。').alignment = WRAP
    ws.cell(n + 2, 1, '加えて、実装状況の確認と画面の撮影が終わるまで公開しないでください。').alignment = WRAP
    for col, w in zip('ABCD', [8, 46, 44, 30]):
        ws.column_dimensions[col].width = w

    wb.save(XLSX)
    print(f'  ✅ {os.path.basename(XLSX)}  本文{wb["ヘルプページ本文"].max_row}行・'
          f'撮影{len(shots)}点・未確定の判断{len(open_j)}件')


if __name__ == '__main__':
    b = md_body()
    build_html(b)
    build_xlsx(b)

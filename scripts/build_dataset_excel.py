#!/usr/bin/env python3
# 20260805_従業員管理_04_検証データセット.xlsx を data/ と検証プランから作り直す。
#
# なぜ自動生成にするか:
#   このブックは data/ のTSV/CSVと、検証プランの準備行（S-）を写したものだった。
#   写しなので、data/ を直しても・検証プランを直しても追随せず、実際に
#   「02_従業員マスタに想定労働日数が無い」「投入手順の画面名が旧名のまま」
#   といった食い違いが起きた。手で直すのをやめ、毎回作り直す。
#
# 使い方:  python3 scripts/build_dataset_excel.py

import csv
import os
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, 'data')
PLAN = os.path.join(ROOT, '20260803_03_従業員管理_検証プラン.tsv')
OUT = os.path.join(ROOT, '20260805_従業員管理_04_検証データセット.xlsx')

# 各ファイルが何のためにあるか。data/ を増やしたらここに1行足す。
PURPOSE = {
    '01_店舗マスタ.tsv': '所属店舗・ヘルプ先店舗の選択肢',
    '02_従業員マスタ.tsv': '一覧・詳細の母数（12名・全項目）',
    '03_ヘルプ先交通費.tsv': 'ヘルプ先交通費の動的フォーム',
    '04_設定変更履歴.tsv': '設定変更履歴（予約中1件を含む・変更ソース4種）',
    '05_月次給与実績.tsv': '月次給与実績（交通費を人件費に含む・5名）',
    '06_日次給与実績.tsv': '日次給与実績（日次合計＝月次）',
    '07_インポートCSV_KingOfTime_sample.csv': 'King of Time 形式の取込',
    '08_インポートCSV_ジョブカン_sample.csv': 'ジョブカン形式の取込',
    '09_インポートCSV_異常系_sample.csv': '取込エラーの検出',
    '10_インポートCSV_給与設定_sample.csv': '給与設定の取込',
    '11_インポートCSV_先頭説明行あり_sample.csv': '先頭に説明行があるCSVの取込',
    '12_従業員マスタ_設定未入力.tsv': '設定が欠落した従業員の扱い',
    '13_勤務データ_2026-05_全稼働日.tsv': '2026-05の全稼働日（5名95行）',
    '14_勤務データ_2026-06_月中改定検証.tsv': '2026-06 月中改定の検証用',
    '15_移行前後_突合テンプレート.tsv': '移行前後の突合（No-Go判定に使う）',
    '16_前回取込_設定_差分確認用.tsv': '前回取込との差分・設定',
    '17_前回取込_実績_差分確認用.tsv': '前回取込との差分・実績',
    '18_列マッピング表_テンプレ_sample.tsv': '列マッピング表テンプレ',
    '19_インポートCSV_重複コード_sample.csv': '同じ従業員コードが2行・二重登録の防止',
    '20_外部ID対応表.tsv': '勤怠システムの外部IDと従業員IDの対応（取込の突き合わせ）',
}

TITLE = Font(bold=True, size=11)
HEAD = Font(bold=True, size=10)
BODY = Font(size=10)
FILL = PatternFill('solid', fgColor='F1F3F5')
LINE = Border(bottom=openpyxl.styles.Side(style='thin', color='CED4DA'))
WRAP = Alignment(vertical='top', wrap_text=True)
TOP = Alignment(vertical='top')


def read(path):
    d = '\t' if path.endswith('.tsv') else ','
    with open(path, encoding='utf-8-sig', newline='') as f:
        return [r for r in csv.reader(f, delimiter=d)]


def uses(num, row):
    """検証プランの1行が data/NN を使っているか。IDのS-01・R-01と混ざらないよう
    「使用する検証データ」列を主に見る。"""
    use = row['使用する検証データ']
    if re.search(rf'(?<![A-Za-z0-9-]){num}(?=[ 　・、）)]|$)', use):
        return True
    blob = use + ' ' + row['操作・前提条件'] + ' ' + row.get('備考', '')
    return bool(re.search(rf'data/{num}(?![0-9])', blob))


def sheet_of(wb, name, title):
    ws = wb.create_sheet(name[:31])
    ws.cell(1, 1, title).font = TITLE
    return ws


def main():
    files = sorted(f for f in os.listdir(DATA) if f.endswith(('.tsv', '.csv')))
    with open(PLAN, encoding='utf-8-sig', newline='') as f:
        plan = list(csv.DictReader(f, delimiter='\t'))

    miss = [f for f in files if f not in PURPOSE]
    if miss:
        raise SystemExit(f'PURPOSE に説明が無いファイル: {miss}\n'
                         f'scripts/build_dataset_excel.py に1行足してください。')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- 00_目次 ----
    ws = sheet_of(wb, '00_目次', f'従業員管理 検証データセット ─ {len(files)}ファイル')
    for i, h in enumerate(['ファイル', '行数', '用途', '使う検証ケース', '件数'], 1):
        c = ws.cell(4, i, h)
        c.font, c.fill, c.border, c.alignment = HEAD, FILL, LINE, TOP
    for n, f in enumerate(files, 5):
        rows = read(os.path.join(DATA, f))
        ids = [r['検証ID'] for r in plan if uses(f[:2], r)]
        vals = [f, len(rows) - 1, PURPOSE[f], ' '.join(ids), len(ids)]
        for i, v in enumerate(vals, 1):
            c = ws.cell(n, i, v)
            c.font, c.alignment = BODY, WRAP
        ws.row_dimensions[n].height = max(14, 13 * (len(' '.join(ids)) // 20 + 1))
    for col, w in zip('ABCDE', [36, 8, 56, 40, 8]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A5'

    # ---- 01_投入手順（検証プランの準備行から作る）----
    prep = [r for r in plan if r['段階'] == '準備']
    ws = sheet_of(wb, '01_投入手順',
                  f'投入手順（検証プランの準備行 {prep[0]["検証ID"]}〜{prep[-1]["検証ID"]}）')
    for i, h in enumerate(['順', 'やること', '投入・確認する内容', '期待される状態'], 1):
        c = ws.cell(3, i, h)
        c.font, c.fill, c.border, c.alignment = HEAD, FILL, LINE, TOP
    for n, r in enumerate(prep, 4):
        for i, v in enumerate([r['検証ID'], r['検証内容'],
                               r['操作・前提条件'], r['期待値']], 1):
            c = ws.cell(n, i, v)
            c.font, c.alignment = BODY, WRAP
        longest = max(len(r['操作・前提条件']) // 34,
                      len(r['期待値']) // 14, 1)
        ws.row_dimensions[n].height = 14 * longest + 4
    for col, w in zip('ABCD', [9, 26, 74, 30]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A4'

    # ---- 各データファイル ----
    for f in files:
        rows = read(os.path.join(DATA, f))
        ws = sheet_of(wb, os.path.splitext(f)[0], f'{f}　─　{PURPOSE[f]}')
        for i, h in enumerate(rows[0], 1):
            c = ws.cell(3, i, h)
            c.font, c.fill, c.border, c.alignment = HEAD, FILL, LINE, TOP
        for n, row in enumerate(rows[1:], 4):
            for i, v in enumerate(row, 1):
                c = ws.cell(n, i, int(v) if re.fullmatch(r'-?\d+', v) else v)
                c.font, c.alignment = BODY, TOP
        for i in range(1, len(rows[0]) + 1):
            width = max(len(str(r[i - 1])) for r in rows if i <= len(r))
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = \
                min(40, max(11, width + 2))
        ws.freeze_panes = 'A4'

    wb.save(OUT)
    print(f'  ✅ {os.path.basename(OUT)}  '
          f'目次1＋投入手順1＋データ{len(files)} ＝ {len(wb.sheetnames)}シート')


if __name__ == '__main__':
    main()
